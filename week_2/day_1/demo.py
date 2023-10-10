from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws1 = wb.create_sheet("Rugrats")
ws3 = wb.create_sheet("Hey Arnold", 0) # insert at first position

for sheet in wb:
    print(sheet.title)

#Accessing one cell
c = ws['A4']


#This provides access to cells using row and column notation:
d = ws.cell(row=4, column=2, value="Whatever I want")

#this will create 100x100 cells in memory
for x in range(1, 101):
    for y in range(1, 101):
        ws.cell(row=x, column=y)

  # a single row or a single column
colC = ws["C"]  
row10 = ws[10]

for cell in colC:
    cell.value ='new data'

for row in ws.iter_rows(min_row=1, max_row=10, min_col=5, max_col=10): # for every row in between 1-10 row and 5-10 (E-J) column
 for cell in row:
    print(cell)

#SAVING A FILE
wb.save("./week_2/spreadsheets/demo.xlsx")    
