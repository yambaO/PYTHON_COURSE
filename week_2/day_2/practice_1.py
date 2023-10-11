# ITP Week 2 Day 2 (In-Class) Practice 1
# 
# You will continue to work on the inventory spreadsheet that you created from yesterday's exercise
# import the appropriate method from the correct module
from openpyxl import load_workbook 


# Import the workbook that you created in yesterday's exercise from
# "./spreadsheets/inventory.xlsx"
wb = load_workbook("/Users/yamba/Python_course/week_2/spreadsheets/inventory.xlsx")

# verify what sheet names are available

for sheet in wb:
    print(sheet.title)

# access and store the appropriate worksheet to the variable 'ws'
ws = wb["CURRENT_MONTH_INVENTORY"]

# Print out the cell values for each row
all_rows = ws.rows
all_values = ws.values
print(list(all_values))

# Create a new column within that worksheet called order_amount
ws["F1"] = "order_amount"
ws.cell(column=6, row=1, value="order_amount")
# save the latest changes
wb.save("/Users/yamba/Python_course/week_2/spreadsheets/inventory.xlsx")