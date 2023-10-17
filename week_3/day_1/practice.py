# ITP Week 3 Day 1 Practice

# import your required modules/methods
from openpyxl import Workbook
wb = Workbook()
ws = wb.active
# given the following items, using the methods we covered, write to openpyxl

# use an external counter with just a for loop (no function)
clefairy = {
    "id": 35,
    "name": "clefairy",
    "base_experience": 113,
    "height": 6,
    "order": 56,
    "weight": 75,
}
count = 1
for key, value in clefairy.items():
    ws.cell(row=1,column=count, value=key)
    ws.cell(row=2,column=count, value=value)
    count+= 1

# create a function that takes in a pokemon
weedle = {
    "id": 13,
    "name": "weedle",
    "base_experience": 39,
    "height": 3,
    "order": 17,
    "weight": 32
}
def pokeball(pokemon):
    c_count = 1
    for key in pokemon:
         ws.cell(row=3, column=c_count, value=pokemon[key])
    c_count+=1
# call the function with weedle!
pokeball(weedle)

wb.save('/Users/yamba/Python_course/week_3/spreadsheets/practice.xlsx')