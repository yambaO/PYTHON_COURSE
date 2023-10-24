# ITP Week 3 Day 3 Exercise

# RICK AND MORTY API DOCS: https://rickandmortyapi.com/documentation

# we want to make a copy of the Rick and Morty database (which is provided through the api)


# EASY MODE
# import the appropriate modules (you have 3)
import requests
import json
from openpyxl import Workbook

# character_url = "https://rickandmortyapi.com/api/character"
# set up a workbook and worksheet titled "Rick and Morty Characters"

wb = Workbook()
ws = wb.active
ws.title = "Rick and Morty Characters"

# # assign a variable 'data' with the returned GET request
character_url = "https://rickandmortyapi.com/api/character"
response = requests.get(character_url)
data = response.json()
print(data)

# create the appropriate headers in openpyxl for all of the keys for a single character
ws.append(["Name", "Status", "Species", "location", "Episode Numbers"])

# for row, character in enumerate(data, 2):
#   for column in enumerate(data, 1):
#     ws.cell(row=row, column=column, value=character.value())

for row, character in enumerate(data['results'], 2):
    ws.cell(row=row, column=1, value=character['name'])
    ws.cell(row=row, column=2, value=character['status'])
    ws.cell(row=row, column=3, value=character['species'])
    ws.cell(row=row, column=4, value=character['location']['name'])
    ws.cell(row=row, column=5, value=len(character['episode']))

wb.save("/Users/yamba/Python_course/week_3/spreadsheets/exercise3.xlsx")

# loop through all of the 'results' of the data to populate the rows and columns for each character

# NOTE: due to the headers, the rows need to be offset by one!

# MEDIUM MODE

# create 2 new worksheets for "Rick and Morty Locations" and "Rick and Morty Episodes"

# create 2 new variables for episode_url and location_url (retrieve it from the docs!)
# "https://rickandmortyapi.com/api/location"
# "https://rickandmortyapi.com/api/episode"

# populate the new worksheets appropriately with all of the data!

# NOTE: don't forget your headers!

# HARD MODE
# Can you decipher the INFO key of the data to use "next" url to continuously pull data?
# Currently, we are only pulling 20 items per api pull!
# WE WANT EVERYTHING. (contact instructors for office hours if stuck!)

# NIGHTMARE
# The inner information for characters, locations, and episodes, references one another through urls
# ie. for episode 28, it lists all the character but by their url
# can you use the URLs to make a subsequent call inside your for loops
# to replace the url with just the appropriate names?
# NOTE: need to make use of if statements to see if url exists or not
# (contact instructors for office hours if stuck!)


# wb.save("./spreadsheets/exercise.xlsx")
